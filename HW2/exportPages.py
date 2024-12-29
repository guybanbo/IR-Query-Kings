import requests
from firebase import firebase
import openpyxl
from openpyxl.utils import get_column_letter

# Initialize Firebase connection
url = "https://query-kings-default-rtdb.europe-west1.firebasedatabase.app/"
FBconn = firebase.FirebaseApplication(url, None)

# Function to fetch data from Firebase
def fetch_data_from_firebase():
    # Get data from Firebase 
    result = FBconn.get('/pages', None)
    if result:
        return result
    else:
        print("No data found in Firebase")
        return {}

# Function to create and save Excel file
def save_links_to_excel(links_data):
    # Create a new workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pages"

    # Set the header row
    ws['A1'] = "Link"
    ws['B1'] = "Title"

    # Write the data to the sheet
    row = 2  # Start from row 2 to leave space for headers
    for link, data in links_data.items():
        ws[f'A{row}'] = data.get('link', 'No Link')
        ws[f'B{row}'] = data.get('title', 'No Title')
        row += 1

    # Expand columns based on content length
    for col in ['A', 'B']:
        max_length = 0
        for cell in ws[col]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = max_length + 2  # Adjust width with a little padding
        ws.column_dimensions[get_column_letter(ws[col][0].column)].width = adjusted_width

    # Save the workbook to a file
    wb.save("pages.xlsx")
    print("Data successfully saved to 'pages.xlsx'")

# Fetch data from Firebase
links_data = fetch_data_from_firebase()

# Save the data to Excel
save_links_to_excel(links_data)
